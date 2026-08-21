"""Read an uploaded BOM (Bill Of Materials) file and match each line against
the product catalog by part number (Viec 13 / CV62, REQ-37/REQ-38).

Accepts .csv or .xlsx/.xls, expects 3 columns in this order: part number,
quantity, note (note is optional). The first row is skipped only if it does
not look like a valid quantity, so both "with header" and "no header" files
are accepted without extra configuration - wholesale customers download
these files from all kinds of ERP/inventory tools and headers are not
consistent.
"""
import csv
import io

from openpyxl import load_workbook

# catalogue owns the products collection; read-only cross-app import, no
# writes happen from here.
from catalogue.db import PRODUCTS, get_db


class BomImportError(Exception):
    pass


def read_bom(uploaded_file):
    """uploaded_file: a Django UploadedFile (from request.FILES).

    Returns a list of {part_number, quantity, note} dicts, quantity already
    an int (defaults to 1 if missing/invalid).
    """
    name = (uploaded_file.name or '').lower()
    if name.endswith('.csv'):
        rows = _read_csv(uploaded_file)
    elif name.endswith('.xlsx') or name.endswith('.xls'):
        rows = _read_xlsx(uploaded_file)
    else:
        raise BomImportError('Unsupported file type. Please upload a .csv or .xlsx file.')

    if not rows:
        raise BomImportError('The file has no data rows.')
    return rows


def _read_csv(uploaded_file):
    text = uploaded_file.read().decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    return _parse_rows(reader)


def _read_xlsx(uploaded_file):
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ([cell.value for cell in row] for row in ws.iter_rows())
    return _parse_rows(rows)


def _parse_rows(rows):
    out = []
    for i, row in enumerate(rows):
        if not row or row[0] in (None, ''):
            continue
        part_number = str(row[0]).strip()
        quantity = _as_int(row[1]) if len(row) > 1 else None
        note = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, '') else ''

        if i == 0 and quantity is None and not part_number.replace('-', '').replace('_', '').isalnum():
            continue  # looks like a header row ("Part Number, Qty, Note"), skip it
        if i == 0 and quantity is None and part_number.lower() in ('part number', 'part_number', 'mpn'):
            continue

        out.append({'part_number': part_number, 'quantity': quantity or 1, 'note': note})
    return out


def _as_int(value):
    try:
        n = int(float(value))
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def match_part_numbers(rows):
    """Attach product_id/matched to every row by looking up the catalog.

    A part number that is not found is kept in the result (matched=False)
    rather than dropped, so the customer can see and fix it before the RFQ
    is submitted (CV62 step 2).
    """
    part_numbers = [r['part_number'] for r in rows]
    found = {
        p['part_number']: p['_id']
        for p in get_db()[PRODUCTS].find(
            {'part_number': {'$in': part_numbers}}, {'part_number': 1}
        )
    }
    for r in rows:
        product_id = found.get(r['part_number'])
        r['product_id'] = product_id
        r['matched'] = product_id is not None
        r['unit_price'] = None  # filled in later by the admin when quoting
    return rows
